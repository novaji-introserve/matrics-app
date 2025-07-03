from odoo import _, api, fields, models
import logging
import base64
import io
from openpyxl import load_workbook
from odoo.exceptions import UserError
from copy import copy
from openpyxl.utils.cell import coordinate_from_string

_logger = logging.getLogger(__name__)


class ReportRuns(models.Model):
    _name = 'res.regulatory.report.run'
    _description = 'Regulatory Report Runs'
    _order = 'create_date desc'
    name = fields.Char(string='Name')
    report_id = fields.Many2one(
        comodel_name='res.regulatory.report', string='Report',index=True)
    # Result fields
    processed_file = fields.Binary(string="Report File", tracking=True)
    processed_filename = fields.Char(
        string="Report Filename", tracking=True)
    changes_count = fields.Integer(
        string="Number of Changes Made", readonly=True)

    def action_submit_report(self):
        return {
            "type": "ir.actions.client",
                    "tag": "display_notification",
                    "params": {
                        "title": "Operation successful",
                        "message": 'Report submitted successfully',
                        "type": "success",
                        "sticky": True,
                    }
        }


class Report(models.Model):
    _name = 'res.regulatory.report'
    _description = 'Regulatory Report'
    _order = 'name, create_date desc'
    _inherit = ['mail.thread', 'mail.activity.mixin']
    name = fields.Char(string='Name', required=True, tracking=True)
    template_id = fields.Many2one(
        comodel_name='res.regulatory.report.template', string='Report Template', required=True, tracking=True,index=True)
    date_from = fields.Date(string='Period Start',
                            required=True, index=True, tracking=True)
    date_to = fields.Date(string='Period End',
                          required=True, index=True, tracking=True)
    run_mode = fields.Selection(string='Run Mode', selection=[('auto', 'Automated'), ('manual', 'Manual')],default='auto',index=True)
    run_frequency = fields.Selection(string='Run Frequency', selection=[('daily', 'Daily'), ('weekly', 'Weekly'),('monthly','Monthly')],default='monthly')
    # Result fields
    processed_file = fields.Binary(string="Processed File", tracking=True)
    processed_filename = fields.Char(
        string="Processed Filename", tracking=True)
    changes_count = fields.Integer(
        string="Number of Changes Made", readonly=True)
    run_ids = fields.One2many(
        'res.regulatory.report.run', 'report_id', string='Report Runs')

    def action_process_report(self):
        """
        Main method to process Excel file and perform find/replace operations
        """
        if not self.template_id.template_file:
            raise UserError("Please upload a template file first.")

        try:
            # Decode binary data

            file_data = base64.b64decode(self.template_id.template_file)

            # Create Excel workbook object from binary data
            workbook = load_workbook(io.BytesIO(file_data))
            changes_count = 0
            # Perform find and replace
            item_ids = self.template_id.item_ids
            if item_ids:
                for i in item_ids:
                    changes_count = self._find_replace_in_workbook(
                        workbook,
                        i.name.strip().upper(),
                        i.item_id.get_value()
                    )
            # Convert back to binary
            processed_binary = self._workbook_to_binary(workbook)
            run = self.env['res.regulatory.report.run'].create({
                'name': self.name,
                'processed_file': processed_binary,
                'processed_filename': f"{self.template_id.entity_id.code}_{self.template_id.code}",
                'changes_count': 1,
                'report_id': self.id
            })

            return {
                'type': 'ir.actions.act_window',
                'name': 'Report Processing Complete',
                'view_mode': 'form',
                'res_model': 'res.regulatory.report',
                'res_id': self.id,
                'target': 'current',
            }

        except Exception as e:
            raise UserError(f"Error processing file: {str(e)}")

    def _find_replace_in_workbook(self, workbook, find_val, replace_val):
        """
        Find and replace values in all worksheets of the workbook

        Args:
            workbook: openpyxl Workbook object
            find_val: Value to find
            replace_val: Value to replace with

        Returns:
            int: Number of replacements made
        """
        changes_count = 0

        # Iterate through all worksheets
        for worksheet in workbook.worksheets:

            if isinstance(replace_val, dict):
                pass
            if isinstance(replace_val, list):
                cell = worksheet[find_val]
                col_letter, row_num = coordinate_from_string(find_val)
                # Get all cells in that row so we can apply styles per cell
                row_cells = []
                for col in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row_num, column=col)
                    row_cells.append(cell)
                start_row = cell.row
                new_data = replace_val
                worksheet.insert_rows(start_row, len(new_data))
                for i, row in enumerate(new_data):
                    list_row = list(row)
                    for j, value in enumerate(list_row):
                        current_col = j+1
                        new_cell = worksheet.cell(
                            row=start_row+i, column=current_col, value=value)
                        copy_cell = row_cells[j]
                        if copy_cell.has_style:
                            new_cell.font = copy(copy_cell.font)
                            new_cell.border = copy(copy_cell.border)
                            new_cell.fill = copy(copy_cell.fill)
            else:
                self._find_replace_in_cell(
                    worksheet, find_val, replace_val
                )
            changes_count = changes_count+1
            return changes_count

    def _find_replace_in_cell(self, worksheet, find_val, replace_val):
        worksheet[f"{find_val.upper()}"] = replace_val

    def _find_replace_in_worksheet(self, worksheet, find_val, replace_val):
        """
        Find and replace values in a specific worksheet

        Args:
            worksheet: openpyxl Worksheet object
            find_val: Value to find
            replace_val: Value to replace with

        Returns:
            int: Number of replacements made in this worksheet
        """
        changes_count = 0

        # Iterate through all cells in the worksheet
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    # Handle different data types
                    cell_value = str(cell.value)

                    # Exact match replacement
                    if cell_value == find_val:
                        cell.value = replace_val
                        changes_count += 1

                    # Partial match replacement (if find_val is substring)
                    elif find_val in cell_value:
                        cell.value = cell_value.replace(find_val, replace_val)
                        changes_count += 1

        return changes_count

    def _workbook_to_binary(self, workbook):
        """
        Convert openpyxl workbook back to binary data

        Args:
            workbook: openpyxl Workbook object

        Returns:
            str: Base64 encoded binary data
        """
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return base64.b64encode(output.read()).decode('utf-8')

    @api.model
    def get_excel_from_record(self, model_name, record_id, field_name):
        """
        Utility method to get Excel binary data from any Odoo record

        Args:
            model_name: Name of the model (e.g., 'res.partner')
            record_id: ID of the record
            field_name: Name of the binary field containing Excel data

        Returns:
            openpyxl Workbook object
        """
        record = self.env[model_name].browse(record_id)
        if not record.exists():
            raise UserError(
                f"Record with ID {record_id} not found in {model_name}")

        binary_data = getattr(record, field_name, None)
        if not binary_data:
            raise UserError(f"No data found in field {field_name}")

        file_data = base64.b64decode(binary_data)
        return load_workbook(io.BytesIO(file_data))

    @api.model
    def batch_find_replace(self, model_name, field_name, find_replace_pairs):
        """
        Batch process multiple records for find/replace operations

        Args:
            model_name: Name of the model containing Excel files
            field_name: Name of the binary field
            find_replace_pairs: List of tuples [(find1, replace1), (find2, replace2), ...]
        """
        records = self.env[model_name].search([(field_name, '!=', False)])

        for record in records:
            try:
                # Get workbook from record
                binary_data = getattr(record, field_name)
                file_data = base64.b64decode(binary_data)
                workbook = load_workbook(io.BytesIO(file_data))

                total_changes = 0

                # Apply all find/replace pairs
                for find_val, replace_val in find_replace_pairs:
                    changes = self._find_replace_in_workbook(
                        workbook, find_val, replace_val)
                    total_changes += changes

                # Save back to record if changes were made
                if total_changes > 0:
                    processed_binary = self._workbook_to_binary(workbook)
                    setattr(record, field_name, processed_binary)

            except Exception as e:
                # Log error but continue with other records
                _logger.error(f"Error processing record {record.id}: {str(e)}")
                continue
