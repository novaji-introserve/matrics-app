import pandas as pd
import numpy as np
import tabula
import os
import logging
import traceback
from datetime import datetime
import xml.etree.ElementTree as ET
import re
import csv
import gc
import uuid
import zipfile
import xml.etree.ElementTree as ET
from io import BytesIO
import subprocess
import pyexcel
import openpyxl
import xlrd


_logger = logging.getLogger(__name__)


class DataProcessor:
    """
    Processes sanctions data files and prepares them for database import
    """

    def __init__(self, env=None):
        """
        Initialize the data processor

        Args:
            env: Odoo environment
        """
        self.env = env

        # Define the batch size for processing
        self.batch_size = 500

        # Define field mappings for different sources
        self.field_mappings = {
            "Nigeria": {
                "Name": "name",
                "First Name": "first_name",
                "Last Name": "surname",
                "Middle Name": "middle_name",
                "Title": "title",
                "Gender": "sex",
                "DOB": "date_of_birth",
                "Date of Birth": "date_of_birth",
                "POB": "place_of_birth",
                "Place of Birth": "place_of_birth",
                "Nationality": "citizenship",
                "Position": "present_position",
                "Address": "official_address",
                "ID Number": "unique_identifier",
                "Status": "status",
                "Remarks": "remarks",
                "Program": "pep_classification",
                "Alias": "aka",
                "Listed On": "createdon",
                "Passport": "additional_info",
            }
        }

        # Fallback generic mapping (used when source-specific mapping isn't available)
        self.generic_mapping = {
            # Person identifiers
            "uid": "unique_identifier",
            "id": "unique_identifier",
            "identifier": "unique_identifier",
            "unique_id": "unique_identifier",
            "person_id": "unique_identifier",
            "group id": "unique_identifier",
            # Name fields
            "surname": "surname",
            "last_name": "surname",
            "lastname": "surname",
            "family_name": "surname",
            "familyname": "surname",
            "name 3": "surname",
            "first_name": "first_name",
            "firstname": "first_name",
            "given_name": "first_name",
            "givenname": "first_name",
            "name 1": "first_name",
            "middle_name": "middle_name",
            "middlename": "middle_name",
            "second_name": "middle_name",
            "name 2": "middle_name",
            "full_name": "name",
            "name": "name",
            "person_name": "name",
            "name 6": "name",
            # Titles and classifications
            "title": "title",
            "position": "present_position",
            "current_position": "present_position",
            "occupation": "present_position",
            "role": "present_position",
            "designation": "present_position",
            "previous_position": "previous_position",
            "former_position": "previous_position",
            "past_position": "previous_position",
            "classification": "pep_classification",
            "pep_type": "pep_classification",
            "peptype": "pep_classification",
            "pep_classification": "pep_classification",
            "regime": "pep_classification",
            "program": "pep_classification",
            "sdntype": "pep_classification",
            "sdn_type": "pep_classification",
            "un_list_type": "pep_classification",
            "group_type": "pep_classification",
            # Personal information
            "alias": "aka",
            "also_known_as": "aka",
            "aka": "aka",
            "alias_name": "aka",
            "gender": "sex",
            "sex": "sex",
            "dob": "date_of_birth",
            "date_of_birth": "date_of_birth",
            "birth_date": "date_of_birth",
            "birthdate": "date_of_birth",
            "dateofbirth": "date_of_birth",
            "pob": "place_of_birth",
            "place_of_birth": "place_of_birth",
            "birth_place": "place_of_birth",
            "birthplace": "place_of_birth",
            "town of birth": "place_of_birth",
            "country of birth": "place_of_birth",
            "placeofbirth": "place_of_birth",
            "nationality": "citizenship",
            "citizenship": "citizenship",
            "citizenships": "citizenship",
            "country": "citizenship",
            "age": "age",
            # Status
            "status": "status",
            "state": "status",
            # Addresses and contact information
            "address": "official_address",
            "official_address": "official_address",
            "business_address": "official_address",
            "address 1": "official_address",
            "city": "official_address",
            "street": "residential_address",
            "residential_address": "residential_address",
            "home_address": "residential_address",
            "address 2": "residential_address",
            "email": "email",
            # Profession and business interests
            "profession": "profession",
            "business": "business_interest",
            "business_interest": "business_interest",
            "business_interests": "business_interest",
            # Additional information fields
            "remarks": "remarks",
            "comments": "remarks",
            "comments1": "remarks",
            "other information": "remarks",
            "note": "remarks",
            # Creation/modification tracking
            "listed_on": "createdon",
            "created_on": "createdon",
            "createdon": "createdon",
            "creation_date": "createdon",
            "created_by": "createdby",
            "createdby": "createdby",
            "last_updated": "lastmodifiedon",
            "modified_on": "lastmodifiedon",
            "lastmodifiedon": "lastmodifiedon",
            "update_date": "lastmodifiedon",
        }

    def _clean_column_names(self, columns):
        """
        Clean and normalize column names

        Args:
            columns: List of column names

        Returns:
            list: Cleaned column names
        """
        cleaned = []
        for col in columns:
            # Convert to string
            col = str(col)

            # Remove extra whitespace
            col = col.strip()

            # Replace line breaks with spaces
            col = col.replace("\n", " ").replace("\r", " ")

            # If column is empty, use a placeholder
            if not col:
                col = "empty_column"

            cleaned.append(col)

        return cleaned

    def _get_mapping(self, source, columns):
        """
        Get field mapping for a specific source

        Args:
            source: Source name (e.g., 'Nigeria')
            columns: List of column names in the data

        Returns:
            dict: Mapping of source columns to model fields
        """
        mapping = {}

        # Get source-specific mapping if available
        source_mapping = self.field_mappings.get(source, self.generic_mapping)

        # Map each column
        for col in columns:
            # Try direct mapping first
            if col in source_mapping:
                mapping[col] = source_mapping[col]
                continue

            # Try case-insensitive mapping
            col_lower = col.lower()
            if col_lower in self.generic_mapping:
                mapping[col] = self.generic_mapping[col_lower]
                continue

            # Try partial matching
            for key, value in self.generic_mapping.items():
                if key in col_lower:
                    mapping[col] = value
                    break

        return mapping

    def _make_unique_identifier(self, row, column_map):
        """
        Generate a unique identifier for a row if none exists

        Args:
            row: Row data
            column_map: Mapping of source columns to model fields

        Returns:
            str: A unique identifier
        """
        # Check if we already have a unique identifier
        for col, field in column_map.items():
            if field == "unique_identifier" and col in row and row[col]:
                return str(row[col])

        # Try to build an identifier from name fields
        name_parts = []

        # Find columns that map to name fields
        name_field_maps = {}
        for col, field in column_map.items():
            if field in ["first_name", "middle_name", "surname", "name"]:
                name_field_maps[field] = col

        # Build name parts in order
        for field in ["first_name", "middle_name", "surname", "name"]:
            if (
                field in name_field_maps
                and name_field_maps[field] in row
                and row[name_field_maps[field]]
            ):
                name_parts.append(str(row[name_field_maps[field]]))

        # If we have name parts, build a unique ID
        if name_parts:
            name_str = "-".join(name_parts)
            # Remove special characters
            name_str = re.sub(r"[^a-zA-Z0-9]", "", name_str)
            # Add timestamp to ensure uniqueness
            timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
            return f"GEN-{name_str}-{timestamp}"

        # Last resort: generate a random ID
        return f"GEN-{uuid.uuid4()}"

    def _format_date(self, date_str):
        """
        Format a date string to a standard format

        Args:
            date_str: Date string in various formats

        Returns:
            str: Standardized date string
        """
        if not date_str or pd.isna(date_str):
            return None

        date_str = str(date_str).strip()

        # Try to parse common date formats
        formats = [
            "%Y-%m-%d",
            "%d/%m/%Y",
            "%m/%d/%Y",
            "%d-%m-%Y",
            "%m-%d-%Y",
            "%Y/%m/%d",
            "%d.%m.%Y",
            "%m.%d.%Y",
            "%Y.%m.%d",
            "%b %d, %Y",
            "%d %b %Y",
            "%B %d, %Y",
            "%d %B %Y",
        ]

        for fmt in formats:
            try:
                # Try to parse the date
                date_obj = datetime.strptime(date_str, fmt)
                # Return in standard format
                return date_obj.strftime("%Y-%m-%d")
            except:
                continue

        # If no format worked, check for partial dates (year only or year and month)
        if re.match(r"^\d{4}$", date_str):
            return f"{date_str}-01-01"  # Year only, assume January 1

        if re.match(r"^\d{4}-\d{1,2}$", date_str):
            return f"{date_str}-01"  # Year and month, assume first day

        # If all else fails, return the original string
        return date_str

    def _clean_value(self, value):
        """
        Clean a data value

        Args:
            value: Value to clean

        Returns:
            object: Cleaned value
        """
        if pd.isna(value):
            return None

        if isinstance(value, (int, float)):
            return value

        value = str(value).strip()

        # Replace line breaks with spaces
        value = value.replace("\n", " ").replace("\r", " ")

        # Replace multiple spaces with single space
        value = re.sub(r"\s+", " ", value)

        # If empty string, return None
        if not value:
            return None

        return value

    def _prepare_record(self, row, column_map, source):
        """
        Prepare a record for insertion into the database with improved validation
        
        Args:
            row: Row data (dict)
            column_map: Mapping of source columns to model fields
            source: Source name
            
        Returns:
            dict: Record ready for insertion
        """
        record = {}

        # Debugging: Log the raw row data
        _logger.debug(f"Raw row data from {source}: {row}")

        # Map each column to its corresponding field
        for col, field in column_map.items():
            if col in row and row[col] is not None:
                value = self._clean_value(row[col])

                # Special handling for certain fields
                if field == "date_of_birth":
                    value = self._format_date(value)
                elif field == "sex" and value:
                    # Normalize gender values
                    value = value.upper()
                    if value in ["M", "MALE"]:
                        value = "Male"
                    elif value in ["F", "FEMALE"]:
                        value = "Female"

                # Verify that value is not None before adding to record
                if value is not None:
                    record[field] = value

        # Add source information
        record["source"] = source

        # Debugging: Log mapped fields
        _logger.debug(f"Mapped fields from {source}: {record.keys()}")

        # Ensure we have a unique identifier
        if "unique_identifier" not in record or not record["unique_identifier"]:
            record["unique_identifier"] = self._make_unique_identifier(row, column_map)
            _logger.debug(f"Generated unique identifier: {record['unique_identifier']}")

        # Ensure we have first_name and surname fields
        self._ensure_name_fields(record)

        # Debugging: Log final record
        _logger.debug(f"Final prepared record: {record}")

        return record

    # def _prepare_record(self, row, column_map, source):
    #     """
    #     Prepare a record for insertion into the database

    #     Args:
    #         row: Row data (dict)
    #         column_map: Mapping of source columns to model fields
    #         source: Source name

    #     Returns:
    #         dict: Record ready for insertion
    #     """
    #     record = {}

    #     # Map each column to its corresponding field
    #     for col, field in column_map.items():
    #         if col in row and row[col] is not None:
    #             value = self._clean_value(row[col])

    #             # Special handling for certain fields
    #             if field == "date_of_birth":
    #                 value = self._format_date(value)
    #             elif field == "sex" and value:
    #                 # Normalize gender values
    #                 value = value.upper()
    #                 if value in ["M", "MALE"]:
    #                     value = "Male"
    #                 elif value in ["F", "FEMALE"]:
    #                     value = "Female"

    #             record[field] = value

    #     # Add source information
    #     record["source"] = source

    #     # Ensure we have a unique identifier
    #     if "unique_identifier" not in record or not record["unique_identifier"]:
    #         record["unique_identifier"] = self._make_unique_identifier(row, column_map)

    #     # Generate name if we have first_name and surname but no name
    #     if "name" not in record and "first_name" in record and "surname" in record:
    #         record["name"] = f"{record['first_name']} {record['surname']}"

    #     # If we have name but no first_name/surname, try to extract them
    #     # if "name" in record and ("first_name" not in record or "surname" not in record):
    #     #     name_parts = record["name"].split()
    #     # If we have name but no first_name/surname, try to extract them
    #     if ("name" in record
    #         and isinstance(record["name"], str)  # Make sure it's actually a string
    #         and ("first_name" not in record or "surname" not in record)):
    #         name_parts = record["name"].split()
    #         if len(name_parts) >= 2:
    #             if "first_name" not in record:
    #                 record["first_name"] = name_parts[0]
    #             if "surname" not in record:
    #                 record["surname"] = name_parts[-1]
    #             if "middle_name" not in record and len(name_parts) > 2:
    #                 record["middle_name"] = " ".join(name_parts[1:-1])
    #     # if "name" in record and record["name"] is not None and ("first_name" not in record or "surname" not in record):
    #     #     name_parts = record["name"].split()
    #     #     if len(name_parts) >= 2:
    #     #         if "first_name" not in record:
    #     #             record["first_name"] = name_parts[0]
    #     #         if "surname" not in record:
    #     #             record["surname"] = name_parts[-1]
    #     #         if "middle_name" not in record and len(name_parts) > 2:
    #     #             record["middle_name"] = " ".join(name_parts[1:-1])

    #     return record

    def _ensure_name_fields(self, record):
        """
        Ensure that the record has the required name fields
        
        Args:
            record: Record to check and fix
            
        Returns:
            None (modifies record in place)
        """
        # Generate name if we have first_name and surname but no name
        if "name" not in record and "first_name" in record and "surname" in record:
            record["name"] = f"{record['first_name']} {record['surname']}"

        # If we have name but no first_name/surname, try to extract them
        if "name" in record and isinstance(record["name"], str):
            name_parts = record["name"].split()
            if len(name_parts) >= 1:
                if "first_name" not in record or not record["first_name"]:
                    record["first_name"] = name_parts[0]

                if len(name_parts) >= 2:
                    if "surname" not in record or not record["surname"]:
                        record["surname"] = name_parts[-1]

                    if "middle_name" not in record and len(name_parts) > 2:
                        record["middle_name"] = " ".join(name_parts[1:-1])

        # Make sure first_name and surname are not empty
        if "first_name" not in record or not record["first_name"]:
            if "name" in record and isinstance(record["name"], str):
                parts = record["name"].split()
                if parts:
                    record["first_name"] = parts[0]
                else:
                    record["first_name"] = "Unknown"
            else:
                record["first_name"] = "Unknown"

        if "surname" not in record or not record["surname"]:
            if "name" in record and isinstance(record["name"], str):
                parts = record["name"].split()
                if len(parts) >= 2:
                    record["surname"] = parts[-1]
                else:
                    record["surname"] = "Unknown"
            else:
                record["surname"] = "Unknown"

    # def _validate_record(self, record):
    #     """
    #     Validate a record before insertion

    #     Args:
    #         record: Record to validate

    #     Returns:
    #         tuple: (is_valid, error_message)
    #     """
    #     # Check required fields
    #     if not record.get("unique_identifier"):
    #         return False, "Missing unique identifier"

    #     # Require at least a name, or first_name and surname
    #     if not record.get("name") and not (
    #         record.get("first_name") and record.get("surname")
    #     ):
    #         return False, "Missing name information"

    #     return True, ""

    def _validate_record(self, record):
        """
        Validate a record before insertion with improved checks
        
        Args:
            record: Record to validate
            
        Returns:
            tuple: (is_valid, error_message)
        """
        # Check required fields
        if not record.get("unique_identifier"):
            return False, "Missing unique identifier"

        # Require at least a first_name and surname
        if not record.get("first_name"):
            return False, "Missing first name"

        if not record.get("surname"):
            return False, "Missing surname"

        # Check string length for key fields to avoid database errors
        max_length = 255  # Typical varchar length in database

        for field in ["unique_identifier", "first_name", "middle_name", "surname", "name"]:
            if field in record and isinstance(record[field], str) and len(record[field]) > max_length:
                # Truncate overly long strings
                record[field] = record[field][:max_length]
                _logger.warning(f"Truncated {field} to {max_length} characters")

        return True, ""

    def process_csv(self, file_path, source_name, callback=None):
        """
        Process a CSV file with fixed pandas errors
        
        Args:
            file_path: Path to the CSV file
            source_name: Name of the source
            callback: Optional callback function to process each batch of records
            
        Returns:
            dict: Processing results
        """
        try:
            _logger.info(f"Processing CSV file: {file_path}")

            # Initialize df as None
            df = None

            # Try first with pandas using default settings
            try:
                _logger.info("Attempting pandas parse with default settings")
                df = pd.read_csv(
                    file_path,
                    encoding='utf-8',
                    engine='c',  # Use the C engine which is faster
                    on_bad_lines='skip',
                    skip_blank_lines=True,
                    dtype=str  # Treat all columns as strings to avoid type errors
                )
                _logger.info("Pandas parse with default settings succeeded")
            except Exception as e:
                _logger.warning(f"Pandas parse with default settings failed: {str(e)}")
                df = None

            # If the first attempt didn't work, try different encodings and delimiters
            if df is None:
                encodings = ["utf-8", "latin-1", "iso-8859-1", "cp1252"]
                delimiters = [',', ';', '\t', '|']

                for encoding in encodings:
                    for delimiter in delimiters:
                        try:
                            _logger.info(f"Trying pandas with {encoding} encoding and '{delimiter}' delimiter")
                            df = pd.read_csv(
                                file_path,
                                encoding=encoding,
                                sep=delimiter,
                                on_bad_lines='skip',
                                skip_blank_lines=True,
                                dtype=str
                            )
                            # Check if we got actual data
                            if not df.empty:
                                _logger.info(f"Successfully parsed with {encoding} encoding and '{delimiter}' delimiter")
                                break
                            else:
                                df = None
                        except Exception as e:
                            df = None
                            continue

                    # Break outer loop if we succeeded
                    if df is not None and not df.empty:
                        break

            # If still failed, try with Python's CSV module
            if df is None:
                try:
                    _logger.info("Falling back to CSV module parsing")
                    # Open with proper encoding detection
                    with open(file_path, 'rb') as f:
                        # Read a sample to detect encoding
                        sample = f.read(4096)

                        # Try to detect encoding
                        try:
                            encoding = 'utf-8'
                            sample.decode('utf-8')
                        except UnicodeDecodeError:
                            try:
                                encoding = 'latin-1'
                                sample.decode('latin-1')
                            except UnicodeDecodeError:
                                encoding = 'cp1252'  # Fallback

                    rows = []
                    with open(file_path, 'r', encoding=encoding, errors='replace') as f:
                        # Try to detect delimiter
                        dialect = csv.Sniffer().sniff(f.read(4096))
                        f.seek(0)

                        reader = csv.reader(f, dialect)
                        headers = next(reader)

                        for row in reader:
                            if len(row) == len(headers):
                                rows.append(dict(zip(headers, row)))

                    # Create DataFrame from rows
                    if rows:
                        df = pd.DataFrame(rows)
                        _logger.info(f"CSV module parsing succeeded with {len(rows)} rows")
                    else:
                        _logger.warning("CSV module parsing found no valid rows")
                        df = None
                except Exception as e:
                    _logger.error(f"CSV module parsing failed: {str(e)}")
                    df = None

            # If all parsing methods failed, create an empty DataFrame with error info
            if df is None:
                _logger.error("All parsing methods failed, creating empty DataFrame")
                df = pd.DataFrame({
                    "error_info": ["Failed to parse file"],
                    "file_path": [file_path],
                    "source": [source_name]
                })

            # Check if the DataFrame has any rows
            if df.empty:
                _logger.warning(f"Parsed DataFrame is empty for file: {file_path}")
                return {
                    "status": "warning",
                    "message": "File was parsed but contained no valid data",
                    "records_processed": 0,
                    "records_valid": 0,
                    "records_invalid": 0,
                }

            # Clean column names
            df.columns = self._clean_column_names(df.columns)

            # Get mapping for this source
            column_map = self._get_mapping(source_name, df.columns)

            if not column_map:
                _logger.warning(f"No field mappings found for CSV file: {file_path}")

            # Process in batches
            records_valid = 0
            records_invalid = 0

            # Calculate batch size based on dataframe size
            # Use smaller batches for larger files to avoid memory issues
            rows_count = len(df)
            if rows_count > 10000:
                batch_size = 100
            elif rows_count > 1000:
                batch_size = 500
            else:
                batch_size = self.batch_size

            _logger.info(f"Processing {rows_count} rows with batch size {batch_size}")

            for i in range(0, len(df), batch_size):
                batch = df.iloc[i : i + batch_size]

                # Convert to dictionaries
                batch_dicts = batch.to_dict("records")

                # Prepare records
                batch_records = []
                invalid_records = []

                for row in batch_dicts:
                    record = self._prepare_record(row, column_map, source_name)

                    # Validate record
                    is_valid, error = self._validate_record(record)

                    if is_valid:
                        batch_records.append(record)
                    else:
                        invalid_records.append((record, error))

                # Update counters
                records_valid += len(batch_records)
                records_invalid += len(invalid_records)

                # Process batch via callback
                if callback and batch_records:
                    callback(batch_records)

                # Log progress
                _logger.info(f"Processed batch of {len(batch)} records from CSV file ({i+1}-{i+len(batch)} of {rows_count})")

            _logger.info(f"Completed processing CSV file: {file_path}")
            _logger.info(f"Valid records: {records_valid}, Invalid records: {records_invalid}")

            return {
                "status": "success",
                "message": f"Processed CSV file: {records_valid} valid records, {records_invalid} invalid",
                "records_processed": records_valid + records_invalid,
                "records_valid": records_valid,
                "records_invalid": records_invalid,
            }

        except Exception as e:
            _logger.error(f"Error processing CSV file: {file_path}")
            _logger.error(f"Error: {str(e)}")
            _logger.error(traceback.format_exc())

            return {
                "status": "error",
                "message": f"Error processing CSV file: {str(e)}",
                "records_processed": 0,
                "records_valid": 0,
                "records_invalid": 0,
            }

    def _preprocess_csv_file(self, file_path):
        """
        Preprocess a CSV file to fix common issues
        
        Args:
            file_path: Path to the CSV file
            
        Returns:
            str: Path to the cleaned file, or None if no cleaning was needed
        """
        try:
            _logger.info(f"Preprocessing CSV file: {file_path}")

            # Check if cleaning is needed
            needs_cleaning = False
            issues_found = []

            # Open in binary mode to check for issues
            with open(file_path, 'rb') as f:
                content = f.read()

                # Check for NULL bytes
                if b'\x00' in content:
                    needs_cleaning = True
                    issues_found.append("NULL bytes")

                # Check for BOM (Byte Order Mark)
                if content.startswith(b'\xef\xbb\xbf'):
                    needs_cleaning = True
                    issues_found.append("BOM marker")

                # Check for non-standard line endings
                if b'\r\r\n' in content:
                    needs_cleaning = True
                    issues_found.append("non-standard line endings")

                # Check for control characters
                control_chars = [bytes([i]) for i in range(1, 32) if i not in [9, 10, 13]]  # Exclude tab, LF, CR
                if any(char in content for char in control_chars):
                    needs_cleaning = True
                    issues_found.append("control characters")

            if not needs_cleaning:
                _logger.info("File doesn't need preprocessing")
                return None

            _logger.info(f"File needs preprocessing. Issues found: {', '.join(issues_found)}")

            # Create a cleaned version of the file
            cleaned_path = f"{file_path}.cleaned"

            with open(file_path, 'rb') as in_file, open(cleaned_path, 'wb') as out_file:
                # Skip BOM if present
                content = in_file.read()
                if content.startswith(b'\xef\xbb\xbf'):
                    content = content[3:]

                # Replace NULL bytes with spaces
                content = content.replace(b'\x00', b' ')

                # Normalize line endings to LF
                content = content.replace(b'\r\r\n', b'\n').replace(b'\r\n', b'\n').replace(b'\r', b'\n')

                # Replace control characters with spaces
                for i in range(1, 32):
                    if i not in [9, 10, 13]:  # tab, LF, CR
                        content = content.replace(bytes([i]), b' ')

                # Remove leading/trailing whitespace from lines
                lines = content.split(b'\n')
                cleaned_lines = [line.strip() for line in lines]

                # Write cleaned content
                out_file.write(b'\n'.join(cleaned_lines))

            _logger.info(f"Created cleaned file: {cleaned_path}")
            return cleaned_path

        except Exception as e:
            _logger.error(f"Error preprocessing CSV file: {str(e)}")
            _logger.error(traceback.format_exc())
            return None

    def process_excel(self, file_path, source_name, callback=None):
        """
        Process an Excel file with format-specific handling
        
        Args:
            file_path: Path to the Excel file
            source_name: Name of the source
            callback: Optional callback function to process each batch of records
            
        Returns:
            dict: Processing results
        """
        try:
            _logger.info(f"Processing Excel file: {file_path}")

            # Determine file type
            _, file_extension = os.path.splitext(file_path)
            file_extension = file_extension.lower()

            # Special handling for ODS files to prevent server crashes
            if file_extension == ".ods":
                _logger.info("ODS file detected, using memory-safe approach")
                return self._safe_process_ods_file(file_path, source_name, callback)

            # For other Excel files, use standard approach with safeguards
            file_size = os.path.getsize(file_path)
            file_size_mb = file_size / (1024 * 1024)

            _logger.info(f"Excel file size: {file_size_mb:.2f} MB")

            # For large files, use extra caution
            if file_size_mb > 5:
                _logger.info("Large Excel file detected, using limited row processing")
                return self._safe_process_excel_large(file_path, source_name, file_extension, callback)

            # Regular processing for small Excel files
            try:
                engine = "openpyxl" if file_extension == ".xlsx" else "xlrd"
                _logger.info(f"Reading Excel with {engine} engine (max 1000 rows)")

                # Read with row limit to prevent memory issues
                df = pd.read_excel(
                    file_path, 
                    engine=engine,
                    nrows=1000,  # Limit to 1000 rows max
                    dtype=str
                )

                _logger.info(f"Successfully read Excel with {len(df)} rows")

                # Clean column names
                df.columns = self._clean_column_names(df.columns)

                # Get mapping for this source
                column_map = self._get_mapping(source_name, df.columns)

                if not column_map:
                    _logger.warning(f"No field mappings found for Excel file: {file_path}")

                # Process in small batches
                records_valid = 0
                records_invalid = 0
                batch_size = 50  # Very small batches

                for i in range(0, len(df), batch_size):
                    batch = df.iloc[i : i + batch_size]

                    # Convert to dictionaries
                    batch_dicts = batch.to_dict("records")

                    # Prepare records
                    batch_records = []
                    invalid_records = []

                    for row in batch_dicts:
                        record = self._prepare_record(row, column_map, source_name)

                        # Validate record
                        is_valid, error = self._validate_record(record)

                        if is_valid:
                            batch_records.append(record)
                        else:
                            invalid_records.append((record, error))

                    # Update counters
                    records_valid += len(batch_records)
                    records_invalid += len(invalid_records)

                    # Process batch via callback
                    if callback and batch_records:
                        callback(batch_records)

                    # Log progress
                    _logger.info(f"Processed batch {i//batch_size + 1}/{(len(df)//batch_size)+1}")

                    # Force garbage collection
                    gc.collect()

                _logger.info(f"Completed processing Excel file: {file_path}")
                _logger.info(f"Valid records: {records_valid}, Invalid records: {records_invalid}")

                return {
                    "status": "success",
                    "message": f"Processed Excel file: {records_valid} valid records, {records_invalid} invalid",
                    "records_processed": records_valid + records_invalid,
                    "records_valid": records_valid,
                    "records_invalid": records_invalid,
                }

            except Exception as e:
                _logger.error(f"Error processing Excel file: {str(e)}")
                return {
                    "status": "error",
                    "message": f"Error processing Excel file: {str(e)}",
                    "records_processed": 0,
                    "records_valid": 0,
                    "records_invalid": 0,
                }

        except Exception as e:
            _logger.error(f"Error in Excel processing: {str(e)}")
            _logger.error(traceback.format_exc())

            return {
                "status": "error",
                "message": f"Error processing Excel file: {str(e)}",
                "records_processed": 0,
                "records_valid": 0,
                "records_invalid": 0,
            }

    def _safe_process_ods_file(self, file_path, source_name, callback=None):
        """
        Process ODS files using extremely memory-safe approaches to prevent server crashes
        
        Args:
            file_path: Path to the ODS file
            source_name: Name of the source
            callback: Optional callback function
            
        Returns:
            dict: Processing results
        """
        _logger.info(f"Using memory-safe ODS processing for: {file_path}")

        # Strategy 1: Try to convert to CSV first (if possible)
        csv_path = self._convert_ods_to_csv(file_path)
        if csv_path:
            _logger.info(f"Successfully converted ODS to CSV, processing the CSV instead")
            return self.process_csv(csv_path, source_name, callback)

        # Strategy 2: Extract a sample of data directly from the ODS without pandas
        try:
            _logger.info("Extracting a limited sample from ODS file")

            # ODS files are ZIP files with XML content
            # Create a dictionary for sample data
            sample_data = []
            headers = []

            # Open the ODS file as a ZIP archive
            with zipfile.ZipFile(file_path, 'r') as z:
                # Check if content.xml exists
                if 'content.xml' not in z.namelist():
                    _logger.error("Invalid ODS file: content.xml not found")
                    return {
                        "status": "warning",
                        "message": "Invalid ODS file structure, skipping",
                        "records_processed": 0,
                        "records_valid": 0,
                        "records_invalid": 0,
                    }

                # Extract just the content.xml file
                content_data = z.read('content.xml')

            # Parse only the beginning of content.xml to extract headers and a few rows
            _logger.info("Parsing ODS header and sample rows")

            # Define ODS namespaces
            namespaces = {
                'office': 'urn:oasis:names:tc:opendocument:xmlns:office:1.0',
                'table': 'urn:oasis:names:tc:opendocument:xmlns:table:1.0',
                'text': 'urn:oasis:names:tc:opendocument:xmlns:text:1.0'
            }

            # Use iterparse to process the file with minimal memory
            context = ET.iterparse(BytesIO(content_data), events=('end',))

            # Extract only the first few rows
            max_rows = 100  # Only process this many rows
            row_count = 0
            in_first_table = False
            current_row = []

            # Process XML in streaming mode
            for event, elem in context:
                if row_count >= max_rows:
                    break

                # Find the first table
                if elem.tag == '{' + namespaces['table'] + '}table' and not in_first_table:
                    in_first_table = True
                    continue

                # Exit if we've moved to another table
                if elem.tag == '{' + namespaces['table'] + '}table' and in_first_table:
                    break

                # Process rows from the first table
                if in_first_table and elem.tag == '{' + namespaces['table'] + '}table-row':
                    # Extract cell values
                    current_row = []
                    for cell in elem.findall('.//{' + namespaces['table'] + '}table-cell'):
                        # Get cell value
                        value = ""
                        for text_elem in cell.findall('.//{' + namespaces['text'] + '}p'):
                            if text_elem.text:
                                value += text_elem.text

                        current_row.append(value)

                    # Process based on row position
                    if row_count == 0:
                        # Header row
                        headers = current_row
                    else:
                        # Data row
                        if len(current_row) > 0:
                            # Make sure we have the right number of columns
                            if len(current_row) < len(headers):
                                current_row.extend([''] * (len(headers) - len(current_row)))
                            elif len(current_row) > len(headers):
                                current_row = current_row[:len(headers)]

                            # Add to sample data
                            row_dict = dict(zip(headers, current_row))
                            sample_data.append(row_dict)

                    # Increment row counter
                    row_count += 1

                    # Clear element to free memory
                    elem.clear()

            # Clean up
            del context

            # Check if we got any valid data
            if not headers or not sample_data:
                _logger.warning("No valid data extracted from ODS file")
                return {
                    "status": "warning",
                    "message": "No valid data found in ODS file (or file structure not supported)",
                    "records_processed": 0,
                    "records_valid": 0,
                    "records_invalid": 0,
                }

            # Clean header names
            headers = self._clean_column_names(headers)

            # Get mapping for this source
            column_map = self._get_mapping(source_name, headers)

            if not column_map:
                _logger.warning(f"No field mappings found for ODS file")

            # Process the sample data
            records_valid = 0
            records_invalid = 0

            # Process all rows at once (small sample)
            batch_records = []
            invalid_records = []

            for row in sample_data:
                record = self._prepare_record(row, column_map, source_name)

                # Validate record
                is_valid, error = self._validate_record(record)

                if is_valid:
                    batch_records.append(record)
                else:
                    invalid_records.append((record, error))

            # Process via callback
            if callback and batch_records:
                callback(batch_records)

            # Update counters
            records_valid = len(batch_records)
            records_invalid = len(invalid_records)

            _logger.info(f"Processed sample of {len(sample_data)} rows from ODS file")
            _logger.info(f"Valid records: {records_valid}, Invalid records: {records_invalid}")
            _logger.warning(f"Note: Only processed first {max_rows} rows of ODS file to prevent memory issues")

            # Force garbage collection
            gc.collect()

            return {
                "status": "success",
                "message": f"Processed ODS sample: {records_valid} valid records, {records_invalid} invalid (limited to {max_rows} rows)",
                "records_processed": records_valid + records_invalid,
                "records_valid": records_valid,
                "records_invalid": records_invalid,
            }

        except Exception as e:
            _logger.error(f"Error in safe ODS processing: {str(e)}")
            _logger.error(traceback.format_exc())

            # Strategy 3: Complete fallback - skip file with informative message
            _logger.warning("All ODS processing methods failed, skipping file to prevent server crash")

            # Create a single placeholder record for this source
            if callback:
                placeholder_record = {
                    "unique_identifier": f"PLACEHOLDER-{source_name}-{int(time.time())}",
                    "first_name": "File",
                    "surname": "Skipped",
                    "name": "ODS File Skipped",
                    "remarks": f"ODS file was too large to process: {os.path.basename(file_path)}",
                    "source": source_name
                }
                callback([placeholder_record])

            return {
                "status": "warning",
                "message": f"ODS file skipped to prevent memory issues: {str(e)}",
                "records_processed": 0,
                "records_valid": 0,
                "records_invalid": 0,
            }

    def _convert_ods_to_csv(self, ods_path):
        """
        Try to convert ODS to CSV using external tools
        
        Args:
            ods_path: Path to the ODS file
            
        Returns:
            str: Path to CSV file if successful, None otherwise
        """
        try:
            _logger.info(f"Attempting to convert ODS to CSV: {ods_path}")
            csv_path = f"{ods_path}.csv"

            # Method 1: Try using LibreOffice if available
            try:                
                # Check if LibreOffice is installed
                lo_paths = ["libreoffice", "soffice", "/Applications/LibreOffice.app/Contents/MacOS/soffice"]
                lo_cmd = None

                for path in lo_paths:
                    try:
                        subprocess.run([path, "--version"], capture_output=True, timeout=5)
                        lo_cmd = path
                        break
                    except:
                        pass

                if lo_cmd:
                    _logger.info(f"Found LibreOffice at {lo_cmd}, using for conversion")
                    # Convert using LibreOffice headless mode
                    out_dir = os.path.dirname(ods_path)
                    cmd = [lo_cmd, "--headless", "--convert-to", "csv", ods_path, "--outdir", out_dir]

                    # Run with timeout
                    process = subprocess.run(cmd, capture_output=True, timeout=120)

                    # Check if file was created
                    base_name = os.path.basename(ods_path)
                    expected_csv = os.path.join(out_dir, os.path.splitext(base_name)[0] + ".csv")

                    if os.path.exists(expected_csv):
                        _logger.info(f"Successfully converted ODS to CSV: {expected_csv}")
                        return expected_csv
            except Exception as e:
                _logger.warning(f"LibreOffice conversion failed: {str(e)}")

            # Method 2: Try using pyexcel if available
            try:                
                _logger.info("Attempting conversion with pyexcel")
                # Use pyexcel for conversion
                pyexcel.save_as(file_name=ods_path, dest_file_name=csv_path)

                if os.path.exists(csv_path):
                    _logger.info(f"Successfully converted ODS to CSV with pyexcel: {csv_path}")
                    return csv_path
            except ImportError:
                _logger.warning("pyexcel not installed, skipping this conversion method")
            except Exception as e:
                _logger.warning(f"pyexcel conversion failed: {str(e)}")

            # If all conversion methods fail
            _logger.warning("All ODS to CSV conversion methods failed")
            return None

        except Exception as e:
            _logger.error(f"Error in ODS to CSV conversion: {str(e)}")
            return None

    def _safe_process_excel_large(self, file_path, source_name, file_extension, callback=None):
        """
        Safely process large Excel files with minimal memory usage
        
        Args:
            file_path: Path to the Excel file
            source_name: Name of the source
            file_extension: File extension (.xlsx or .xls)
            callback: Optional callback function
            
        Returns:
            dict: Processing results
        """
        try:
            _logger.info(f"Using memory-safe approach for large Excel file: {file_path}")

            # For XLSX files, we can use openpyxl in read-only mode
            if file_extension == ".xlsx":
                _logger.info("Processing XLSX with openpyxl in read-only mode")

                # Use read-only mode to minimize memory usage
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

                # Only process the first sheet
                if wb.sheetnames:
                    sheet = wb[wb.sheetnames[0]]
                    _logger.info(f"Processing first sheet: {wb.sheetnames[0]}")

                    # Read headers
                    headers = []
                    for cell in next(sheet.iter_rows()):
                        headers.append(str(cell.value) if cell.value is not None else "")

                    # Clean headers
                    headers = self._clean_column_names(headers)

                    # Get mapping
                    column_map = self._get_mapping(source_name, headers)

                    if not column_map:
                        _logger.warning(f"No field mappings found for Excel file")

                    # Process rows in small batches
                    records_valid = 0
                    records_invalid = 0
                    rows_processed = 0
                    max_rows = 500  # Limit to 500 rows to prevent memory issues

                    # Process rows in batches
                    current_batch = []

                    # Skip the header row
                    rows = sheet.iter_rows(min_row=2)

                    for row in rows:
                        # Stop if we've reached the maximum row limit
                        if rows_processed >= max_rows:
                            break

                        # Create a dictionary for the row
                        row_dict = {}
                        for i, cell in enumerate(row):
                            if i < len(headers):
                                row_dict[headers[i]] = str(cell.value) if cell.value is not None else ""

                        current_batch.append(row_dict)
                        rows_processed += 1

                        # Process batch when it reaches size or at the end
                        if len(current_batch) >= 50 or rows_processed >= max_rows:
                            # Process this batch
                            batch_records = []
                            invalid_records = []

                            for row_data in current_batch:
                                record = self._prepare_record(row_data, column_map, source_name)

                                # Validate record
                                is_valid, error = self._validate_record(record)

                                if is_valid:
                                    batch_records.append(record)
                                else:
                                    invalid_records.append((record, error))

                            # Update counters
                            records_valid += len(batch_records)
                            records_invalid += len(invalid_records)

                            # Process batch via callback
                            if callback and batch_records:
                                callback(batch_records)

                            # Clear batch
                            current_batch = []

                            # Force garbage collection
                            gc.collect()

                    # Close workbook to free memory
                    wb.close()

                    _logger.info(f"Completed processing {rows_processed} rows from Excel file")
                    _logger.info(f"Valid records: {records_valid}, Invalid records: {records_invalid}")

                    return {
                        "status": "success",
                        "message": f"Processed Excel file: {records_valid} valid records, {records_invalid} invalid (limited to {rows_processed} rows)",
                        "records_processed": records_valid + records_invalid,
                        "records_valid": records_valid,
                        "records_invalid": records_invalid,
                    }

            # For XLS files, use xlrd with limits
            elif file_extension == ".xls":                
                _logger.info("Processing XLS with xlrd with row limits")

                # Open the workbook
                wb = xlrd.open_workbook(file_path, on_demand=True)

                # Only process the first sheet
                if wb.sheet_names():
                    sheet = wb.sheet_by_index(0)
                    _logger.info(f"Processing first sheet: {wb.sheet_names()[0]}")

                    # Get row and column counts
                    num_rows = min(500, sheet.nrows)  # Limit to 500 rows
                    num_cols = sheet.ncols

                    # Read headers from first row
                    headers = [str(sheet.cell_value(0, i)) for i in range(num_cols)]

                    # Clean headers
                    headers = self._clean_column_names(headers)

                    # Get mapping
                    column_map = self._get_mapping(source_name, headers)

                    if not column_map:
                        _logger.warning(f"No field mappings found for Excel file")

                    # Process rows in small batches
                    records_valid = 0
                    records_invalid = 0

                    # Process in batches of 50 rows
                    for start_row in range(1, num_rows, 50):
                        end_row = min(start_row + 50, num_rows)

                        batch_records = []
                        invalid_records = []

                        # Process this batch of rows
                        for row_idx in range(start_row, end_row):
                            # Create a dictionary for the row
                            row_dict = {}
                            for col_idx in range(num_cols):
                                if col_idx < len(headers):
                                    row_dict[headers[col_idx]] = str(sheet.cell_value(row_idx, col_idx))

                            # Prepare and validate record
                            record = self._prepare_record(row_dict, column_map, source_name)
                            is_valid, error = self._validate_record(record)

                            if is_valid:
                                batch_records.append(record)
                            else:
                                invalid_records.append((record, error))

                        # Update counters
                        records_valid += len(batch_records)
                        records_invalid += len(invalid_records)

                        # Process batch via callback
                        if callback and batch_records:
                            callback(batch_records)

                        # Force garbage collection
                        gc.collect()

                    # Close workbook to free memory
                    wb.release_resources()

                    _logger.info(f"Completed processing {num_rows-1} rows from Excel file")
                    _logger.info(f"Valid records: {records_valid}, Invalid records: {records_invalid}")

                    return {
                        "status": "success",
                        "message": f"Processed Excel file: {records_valid} valid records, {records_invalid} invalid (limited to {num_rows-1} rows)",
                        "records_processed": records_valid + records_invalid,
                        "records_valid": records_valid,
                        "records_invalid": records_invalid,
                    }

            # Fallback to pandas with strict limits
            _logger.info("Falling back to pandas with strict row limits")

            # Use pandas with strict row limit
            df = pd.read_excel(
                file_path,
                nrows=100,  # Very strict limit
                dtype=str
            )

            # Clean column names
            df.columns = self._clean_column_names(df.columns)

            # Get mapping
            column_map = self._get_mapping(source_name, df.columns)

            # Process all rows at once (small number)
            batch_records = []
            invalid_records = []

            for _, row in df.iterrows():
                record = self._prepare_record(row.to_dict(), column_map, source_name)

                # Validate record
                is_valid, error = self._validate_record(record)

                if is_valid:
                    batch_records.append(record)
                else:
                    invalid_records.append((record, error))

            # Process via callback
            if callback and batch_records:
                callback(batch_records)

            records_valid = len(batch_records)
            records_invalid = len(invalid_records)

            _logger.info(f"Processed {len(df)} rows using pandas fallback")
            _logger.info(f"Valid records: {records_valid}, Invalid records: {records_invalid}")

            return {
                "status": "success",
                "message": f"Processed Excel file: {records_valid} valid records, {records_invalid} invalid (limited to {len(df)} rows)",
                "records_processed": records_valid + records_invalid,
                "records_valid": records_valid,
                "records_invalid": records_invalid,
            }

        except Exception as e:
            _logger.error(f"Error in safe Excel processing: {str(e)}")
            _logger.error(traceback.format_exc())

            # Create a placeholder record
            if callback:
                placeholder_record = {
                    "unique_identifier": f"PLACEHOLDER-{source_name}-{int(time.time())}",
                    "first_name": "File",
                    "surname": "Limited",
                    "name": "Excel File Limited",
                    "remarks": f"Excel file was too large to process fully: {os.path.basename(file_path)}",
                    "source": source_name
                }
                callback([placeholder_record])

            return {
                "status": "warning",
                "message": f"Excel file processing limited to prevent memory issues: {str(e)}",
                "records_processed": 1,
                "records_valid": 1,
                "records_invalid": 0,
            }

    def process_pdf(self, file_path, source_name, callback=None):
        """
        Process a PDF file

        Args:
            file_path: Path to the PDF file
            source_name: Name of the source
            callback: Optional callback function to process each batch of records

        Returns:
            dict: Processing results
        """
        try:
            _logger.info(f"Processing PDF file: {file_path}")

            # Extract tables from PDF
            tables = tabula.read_pdf(file_path, pages="all", multiple_tables=True)

            if not tables:
                _logger.warning(f"No tables found in PDF file: {file_path}")
                return {
                    "status": "warning",
                    "message": "No tables found in PDF file",
                    "records_processed": 0,
                    "records_valid": 0,
                    "records_invalid": 0,
                }

            records_valid = 0
            records_invalid = 0

            # Process each table
            for i, df in enumerate(tables):
                if df.empty:
                    continue

                # Clean column names
                df.columns = self._clean_column_names(df.columns)

                # Get mapping for this source
                column_map = self._get_mapping(source_name, df.columns)

                if not column_map:
                    _logger.warning(
                        f"No field mappings found for PDF table {i+1} in file: {file_path}"
                    )

                # Process in batches
                for j in range(0, len(df), self.batch_size):
                    batch = df.iloc[j : j + self.batch_size]

                    # Convert to dictionaries
                    batch_dicts = batch.to_dict("records")

                    # Prepare records
                    batch_records = []
                    invalid_records = []

                    for row in batch_dicts:
                        record = self._prepare_record(
                            row, column_map, f"{source_name} (PDF table {i+1})"
                        )

                        # Validate record
                        is_valid, error = self._validate_record(record)

                        if is_valid:
                            batch_records.append(record)
                        else:
                            invalid_records.append((record, error))

                    # Update counters
                    records_valid += len(batch_records)
                    records_invalid += len(invalid_records)

                    # Process batch via callback
                    if callback and batch_records:
                        callback(batch_records)

                    # Log progress
                    _logger.info(
                        f"Processed batch of {len(batch)} records from PDF table {i+1} in file: {file_path}"
                    )

            _logger.info(f"Completed processing PDF file: {file_path}")
            _logger.info(
                f"Valid records: {records_valid}, Invalid records: {records_invalid}"
            )

            return {
                "status": "success",
                "message": f"Processed PDF file: {records_valid} valid records, {records_invalid} invalid",
                "records_processed": records_valid + records_invalid,
                "records_valid": records_valid,
                "records_invalid": records_invalid,
            }

        except Exception as e:
            _logger.error(f"Error processing PDF file: {file_path}")
            _logger.error(f"Error: {str(e)}")
            _logger.error(traceback.format_exc())

            return {
                "status": "error",
                "message": f"Error processing PDF file: {str(e)}",
                "records_processed": 0,
                "records_valid": 0,
                "records_invalid": 0,
            }

    def process_xml(self, file_path, source_name, callback=None):
        """
        Process an XML file

        Args:
            file_path: Path to the XML file
            source_name: Name of the source
            callback: Optional callback function to process each batch of records

        Returns:
            dict: Processing results
        """
        try:
            _logger.info(f"Processing XML file: {file_path}")

            # Parse the XML file
            tree = ET.parse(file_path)
            root = tree.getroot()

            # Function to recursively extract text from an element
            def get_element_text(element):
                text = element.text or ""
                for child in element:
                    text += " " + get_element_text(child)
                return text.strip()

            # Find all person/entity elements
            # This is very source-specific, so we try several common patterns
            person_elements = []

            # Try different tag patterns commonly used in sanctions XML files
            for pattern in [
                "sdnEntry",
                "record",
                "entity",
                "individual",
                "person",
                "party",
                "target",
            ]:
                elements = root.findall(f".//{pattern}")
                if elements:
                    person_elements.extend(elements)

                # Try with namespace
                for ns in root.nsmap.values() if hasattr(root, "nsmap") else []:
                    elements = root.findall(f".//{{{ns}}}{pattern}")
                    if elements:
                        person_elements.extend(elements)

            if not person_elements:
                _logger.warning(
                    f"No person/entity elements found in XML file: {file_path}"
                )
                return {
                    "status": "warning",
                    "message": "No person/entity elements found in XML file",
                    "records_processed": 0,
                    "records_valid": 0,
                    "records_invalid": 0,
                }

            records_valid = 0
            records_invalid = 0

            # Process in batches
            batch_records = []
            invalid_records = []

            for i, element in enumerate(person_elements):
                # Extract all fields from the element
                record = {}

                # Extract element attributes
                for attr_name, attr_value in element.attrib.items():
                    # Clean up attribute name (remove namespace)
                    if "}" in attr_name:
                        attr_name = attr_name.split("}", 1)[1]
                    record[attr_name] = attr_value

                # Extract child elements
                for child in element.findall(".//*"):
                    # Get tag name (remove namespace)
                    tag = child.tag
                    if "}" in tag:
                        tag = tag.split("}", 1)[1]

                    # Get element text
                    text = get_element_text(child)

                    if text:
                        record[tag] = text

                # Map fields using source-specific or generic mapping
                column_map = self._get_mapping(source_name, record.keys())

                if not column_map:
                    _logger.warning(
                        f"No field mappings found for XML element {i+1} in file: {file_path}"
                    )

                prepared_record = self._prepare_record(record, column_map, source_name)

                # Validate record
                is_valid, error = self._validate_record(prepared_record)

                if is_valid:
                    batch_records.append(prepared_record)
                else:
                    invalid_records.append((prepared_record, error))

                # Process batch if needed
                if len(batch_records) >= self.batch_size:
                    # Update counters
                    records_valid += len(batch_records)
                    records_invalid += len(invalid_records)

                    # Process batch via callback
                    if callback and batch_records:
                        callback(batch_records)

                    # Log progress
                    _logger.info(
                        f"Processed batch of {len(batch_records)} records from XML file: {file_path}"
                    )

                    # Reset batch
                    batch_records = []
                    invalid_records = []

            # Process remaining records
            if batch_records:
                # Update counters
                records_valid += len(batch_records)
                records_invalid += len(invalid_records)

                # Process batch via callback
                if callback and batch_records:
                    callback(batch_records)

            _logger.info(f"Completed processing XML file: {file_path}")
            _logger.info(
                f"Valid records: {records_valid}, Invalid records: {records_invalid}"
            )

            return {
                "status": "success",
                "message": f"Processed XML file: {records_valid} valid records, {records_invalid} invalid",
                "records_processed": records_valid + records_invalid,
                "records_valid": records_valid,
                "records_invalid": records_invalid,
            }

        except Exception as e:
            _logger.error(f"Error processing XML file: {file_path}")
            _logger.error(f"Error: {str(e)}")
            _logger.error(traceback.format_exc())

            return {
                "status": "error",
                "message": f"Error processing XML file: {str(e)}",
                "records_processed": 0,
                "records_valid": 0,
                "records_invalid": 0,
            }

    def process_file(self, file_info, callback=None):
        """
        Process a file based on its type

        Args:
            file_info: Dictionary with file information
            callback: Optional callback function to process each batch of records

        Returns:
            dict: Processing results
        """
        file_path = file_info["path"]
        file_type = file_info["type"]
        source_name = file_info["source"]

        _logger.info(
            f"Processing file: {file_path} (type: {file_type}, source: {source_name})"
        )

        # Process based on file type
        if file_type == "csv":
            return self.process_csv(file_path, source_name, callback)
        elif file_type in ["xlsx", "xls", "ods"]:
            return self.process_excel(file_path, source_name, callback)
        elif file_type == "pdf":
            return self.process_pdf(file_path, source_name, callback)
        elif file_type == "xml":
            return self.process_xml(file_path, source_name, callback)
        else:
            _logger.warning(f"Unsupported file type: {file_type}")
            return {
                "status": "error",
                "message": f"Unsupported file type: {file_type}",
                "records_processed": 0,
                "records_valid": 0,
                "records_invalid": 0,
            }
